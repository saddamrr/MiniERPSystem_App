# pages/report_page.py
from PySide6.QtWidgets import *
from PySide6.QtCore import *
from PySide6.QtGui import *
from datetime import datetime, timedelta
from database_helper import DatabaseHelper
from utils import safe_float, format_rupiah, safe_int
from styles import STYLE_REFRESH_BUTTON
import calendar
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from windows.transaction_detail_window import TransactionDetailWindow

class ReportPage(QWidget):
    def __init__(self):
        super().__init__()
        self.current_periode = 'harian'
        self.current_date = datetime.now()
        self.transactions = []
        self.filtered_transactions = []
        self.customers = []
        self.products = []
        self.init_ui()
        self.load_report()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(20)
        
        # ==================== HEADER ====================
        header = QHBoxLayout()
        
        title = QLabel("📊 Laporan Penjualan")
        title.setStyleSheet("font-size: 24px; font-weight: 600; color: #0f172a;")
        header.addWidget(title)
        
        header.addStretch()
        
        # Periode selector
        header.addWidget(QLabel("Periode:"))
        self.period_combo = QComboBox()
        self.period_combo.addItems(["Harian", "Bulanan", "Tahunan", "Kustom"])
        self.period_combo.setMinimumHeight(35)
        self.period_combo.setFixedWidth(120)
        self.period_combo.currentTextChanged.connect(self.on_period_change)
        header.addWidget(self.period_combo)
        
        # Date range untuk kustom
        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDate(QDate.currentDate().addDays(-30))
        self.start_date.setFixedWidth(110)
        self.start_date.setMinimumHeight(35)
        self.start_date.setVisible(False)
        
        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDate(QDate.currentDate())
        self.end_date.setFixedWidth(110)
        self.end_date.setMinimumHeight(35)
        self.end_date.setVisible(False)
        
        header.addWidget(QLabel("Dari:"))
        header.addWidget(self.start_date)
        header.addWidget(QLabel("Sampai:"))
        header.addWidget(self.end_date)
        
        # Tombol filter
        filter_btn = QPushButton("Tampilkan")
        filter_btn.setMinimumHeight(35)
        filter_btn.setFixedWidth(100)
        filter_btn.clicked.connect(self.load_report)
        filter_btn.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #2563eb;
            }
        """)
        header.addWidget(filter_btn)
        
        # Tombol export
        export_btn = QPushButton("📎 Export")
        export_btn.setMinimumHeight(35)
        export_btn.setFixedWidth(100)
        export_btn.clicked.connect(self.export_to_excel)
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #f1f5f9;
                color: #475569;
                border: none;
                border-radius: 8px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #e2e8f0;
            }
        """)
        header.addWidget(export_btn)
        
        layout.addLayout(header)
        
        # ==================== SUMMARY CARDS ====================
        summary_group = QGroupBox("Ringkasan")
        summary_group.setStyleSheet("""
            QGroupBox {
                font-weight: 600;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
                margin-top: 8px;
                padding-top: 8px;
                background-color: white;
            }
            QGroupBox::title {
                color: #3b82f6;
                left: 12px;
                padding: 0 8px;
            }
        """)
        self.summary_layout = QHBoxLayout()
        self.summary_layout.setSpacing(15)
        summary_group.setLayout(self.summary_layout)
        layout.addWidget(summary_group)
        
        # ==================== TAB VIEW ====================
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #e2e8f0;
                border-radius: 12px;
                background-color: white;
            }
            QTabBar::tab {
                padding: 10px 20px;
                font-size: 13px;
                font-weight: 500;
            }
            QTabBar::tab:selected {
                background-color: #3b82f6;
                color: white;
            }
        """)
        
        # Tab 1: Detail Transaksi
        self.trans_tab = self.create_transaction_tab()
        self.tab_widget.addTab(self.trans_tab, "📋 Detail Transaksi")
        
        # Tab 2: Penjualan per Brand
        self.brand_tab = self.create_brand_tab()
        self.tab_widget.addTab(self.brand_tab, "🏷️ Penjualan per Brand")
        
        # Tab 3: Laporan Pelanggan
        self.customer_tab = self.create_customer_tab()
        self.tab_widget.addTab(self.customer_tab, "👥 Laporan Pelanggan")
        
        # Tab 4: Laporan Stok
        self.stock_tab = self.create_stock_tab()
        self.tab_widget.addTab(self.stock_tab, "📦 Laporan Stok")

        # Tab 5: Log Stok (BARU)
        self.stock_log_tab = self.create_stock_log_tab()
        self.tab_widget.addTab(self.stock_log_tab, "📋 Log Stok")
        
        layout.addWidget(self.tab_widget)
        
        self.setLayout(layout)
    
    def create_transaction_tab(self):
        """Tab untuk detail transaksi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        
        self.trans_table = QTableWidget()
        self.trans_table.setColumnCount(7)  # Tambah 1 kolom untuk pelanggan
        self.trans_table.setHorizontalHeaderLabels(["No Invoice", "Tanggal", "Kasir", "Pelanggan", "Karung", "KG", "Total"])
        self.trans_table.horizontalHeader().setStretchLastSection(True)
        self.trans_table.setAlternatingRowColors(True)
        self.trans_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.trans_table.setSelectionMode(QTableWidget.SingleSelection)
        
        # Set column widths
        self.trans_table.setColumnWidth(0, 150)  # No Invoice
        self.trans_table.setColumnWidth(1, 150)  # Tanggal
        self.trans_table.setColumnWidth(2, 100)  # Kasir
        self.trans_table.setColumnWidth(3, 150)  # Pelanggan
        self.trans_table.setColumnWidth(4, 80)   # Karung
        self.trans_table.setColumnWidth(5, 80)   # KG
        self.trans_table.setColumnWidth(6, 120)  # Total
        
        # Row height
        self.trans_table.verticalHeader().setDefaultSectionSize(40)
        self.trans_table.verticalHeader().setVisible(False)
        
        # Connect double click event
        self.trans_table.itemDoubleClicked.connect(self.on_transaction_double_click)
        
        layout.addWidget(self.trans_table)
        
        # Info
        info_label = QLabel("💡 Tips: Double klik pada baris untuk melihat detail transaksi")
        info_label.setStyleSheet("color: #64748b; font-size: 11px; margin-top: 8px;")
        layout.addWidget(info_label)
        
        widget.setLayout(layout)
        return widget

    def on_transaction_double_click(self, item):
        """Handle double click on transaction table"""
        row = item.row()
        if row < len(self.filtered_transactions):
            transaction = self.filtered_transactions[row]
            # Gunakan method yang sudah ada
            self.show_transaction_detail(transaction)

    def show_transaction_detail(self, transaction):
        """Show transaction detail dialog - pastikan hanya satu dialog"""
        # Cek apakah dialog sudah ada
        if hasattr(self, '_detail_dialog') and self._detail_dialog and self._detail_dialog.isVisible():
            # Jika sudah ada, bring to front
            self._detail_dialog.raise_()
            self._detail_dialog.activateWindow()
            return
        
        # Ambil detail lengkap transaksi dari database
        trans_id = transaction.get('id_transaksi')
        if trans_id:
            detail_trans = DatabaseHelper.get_transaction_by_id(trans_id)
            if detail_trans:
                # Tambahkan informasi pelanggan jika ada
                customer_id = transaction.get('id_pelanggan')
                if customer_id:
                    customer = DatabaseHelper.get_customer_by_id(customer_id)
                    if customer:
                        detail_trans['pelanggan'] = customer
                        detail_trans['id_pelanggan'] = customer_id
                self._detail_dialog = TransactionDetailWindow(detail_trans, self)
                self._detail_dialog.finished.connect(self.on_detail_dialog_closed)
                self._detail_dialog.exec()
            else:
                self._detail_dialog = TransactionDetailWindow(transaction, self)
                self._detail_dialog.finished.connect(self.on_detail_dialog_closed)
                self._detail_dialog.exec()
        else:
            self._detail_dialog = TransactionDetailWindow(transaction, self)
            self._detail_dialog.finished.connect(self.on_detail_dialog_closed)
            self._detail_dialog.exec()

    def on_detail_dialog_closed(self):
        """Called when detail dialog is closed"""
        self._detail_dialog = None
    
    def create_brand_tab(self):
        """Tab untuk penjualan per brand"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        
        self.brand_table = QTableWidget()
        self.brand_table.setColumnCount(5)
        self.brand_table.setHorizontalHeaderLabels(["Brand", "Transaksi", "Karung", "KG", "Total Penjualan"])
        self.brand_table.horizontalHeader().setStretchLastSection(True)
        self.brand_table.setAlternatingRowColors(True)
        
        self.brand_table.setColumnWidth(0, 150)
        self.brand_table.setColumnWidth(1, 100)
        self.brand_table.setColumnWidth(2, 100)
        self.brand_table.setColumnWidth(3, 100)
        self.brand_table.setColumnWidth(4, 150)
        
        self.brand_table.verticalHeader().setDefaultSectionSize(40)
        self.brand_table.verticalHeader().setVisible(False)
        
        layout.addWidget(self.brand_table)
        widget.setLayout(layout)
        return widget
    
    def create_customer_tab(self):
        """Tab untuk laporan pelanggan dengan double click untuk history"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(15)
        
        # Filter bar
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Cari Pelanggan:"))
        self.customer_search_input = QLineEdit()
        self.customer_search_input.setPlaceholderText("Nama atau kode pelanggan...")
        self.customer_search_input.setMinimumHeight(32)
        self.customer_search_input.setFixedWidth(250)
        self.customer_search_input.textChanged.connect(self.filter_customer_report)
        filter_layout.addWidget(self.customer_search_input)
        
        filter_layout.addStretch()
        
        # Tombol refresh
        refresh_btn = QPushButton("🔄 Refresh")
        refresh_btn.setMinimumHeight(32)
        refresh_btn.setFixedWidth(100)
        refresh_btn.clicked.connect(self.update_customer_tab)
        filter_layout.addWidget(refresh_btn)
        
        layout.addLayout(filter_layout)
        
        # ==================== TABEL PELANGGAN ====================
        # Gunakan QTableWidget langsung (scroll otomatis)
        self.customer_table = QTableWidget()
        self.customer_table.setColumnCount(4)
        self.customer_table.setHorizontalHeaderLabels(["Kode", "Nama", "Total Belanja", "Transaksi"])
        self.customer_table.setAlternatingRowColors(True)
        self.customer_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.customer_table.setSelectionMode(QTableWidget.SingleSelection)
        
        # Set column widths
        self.customer_table.setColumnWidth(0, 100)
        self.customer_table.setColumnWidth(1, 200)
        self.customer_table.setColumnWidth(2, 150)
        self.customer_table.setColumnWidth(3, 100)
        
        self.customer_table.verticalHeader().setDefaultSectionSize(40)
        self.customer_table.verticalHeader().setVisible(False)
        
        # Connect DOUBLE CLICK event
        self.customer_table.itemDoubleClicked.connect(self.on_customer_table_double_click)
        
        layout.addWidget(self.customer_table)
        
        # Info
        info_label = QLabel("💡 Tips: Double klik pada baris pelanggan untuk melihat history transaksi lengkap")
        info_label.setStyleSheet("color: #64748b; font-size: 11px; margin-top: 8px;")
        layout.addWidget(info_label)
        
        widget.setLayout(layout)
        return widget
    
    def on_customer_selected(self):
        """Handle when a customer is selected in the table"""
        selected_rows = self.customer_table.selectedItems()
        if not selected_rows:
            return
        
        row = selected_rows[0].row()
        
        # Dapatkan data dari tabel yang sedang ditampilkan (bisa filtered)
        current_data = []
        for i in range(self.customer_table.rowCount()):
            kode = self.customer_table.item(i, 0).text() if self.customer_table.item(i, 0) else ""
            nama = self.customer_table.item(i, 1).text() if self.customer_table.item(i, 1) else ""
            current_data.append({'kode': kode, 'nama': nama})
        
        if row < len(current_data):
            customer_kode = current_data[row]['kode']
            customer_nama = current_data[row]['nama']
            
            # Cari di all_customer_data berdasarkan kode atau nama
            customer = None
            if hasattr(self, 'all_customer_data'):
                for c in self.all_customer_data:
                    if c.get('kode_pelanggan') == customer_kode or c.get('nama') == customer_nama:
                        customer = c
                        break
            
            if customer:
                # Update info label
                # self.selected_customer_name.setText(f"👤 {customer.get('nama')} ({customer.get('kode_pelanggan')})")
                
                # Show purchase details
                self.show_customer_purchase_details(customer)
    
    def create_stock_tab(self):
        """Tab untuk laporan stok"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        
        self.stock_table = QTableWidget()
        self.stock_table.setColumnCount(6)
        self.stock_table.setHorizontalHeaderLabels(["Kode", "Brand", "Stok (Karung)", "Stok (KG)", "Harga/KG", "Status"])
        self.stock_table.horizontalHeader().setStretchLastSection(True)
        self.stock_table.setAlternatingRowColors(True)
        
        self.stock_table.setColumnWidth(0, 90)
        self.stock_table.setColumnWidth(1, 150)
        self.stock_table.setColumnWidth(2, 100)
        self.stock_table.setColumnWidth(3, 100)
        self.stock_table.setColumnWidth(4, 100)
        self.stock_table.setColumnWidth(5, 100)
        
        self.stock_table.verticalHeader().setDefaultSectionSize(40)
        self.stock_table.verticalHeader().setVisible(False)
        
        layout.addWidget(self.stock_table)
        widget.setLayout(layout)
        return widget
    
    def update_customer_tab(self):
        """Update customer report table with purchase details"""
        try:
            self.customers = DatabaseHelper.get_customers()
            
            if not self.customers:
                self.customer_table.setRowCount(0)
                self.customer_table.setHorizontalHeaderLabels(["Kode", "Nama", "Total Belanja", "Total Transaksi", "Terakhir Transaksi"])
                return
            
            # Ambil semua transaksi untuk menghitung detail per pelanggan
            all_transactions = DatabaseHelper.get_transactions_with_details()
            
            # Hitung statistik per pelanggan
            customer_stats = {}
            customer_purchases = {}  # Menyimpan detail pembelian per pelanggan
            
            for t in all_transactions:
                customer_id = t.get('id_pelanggan')
                if customer_id:
                    if customer_id not in customer_stats:
                        customer_stats[customer_id] = {
                            'total_belanja': 0,
                            'total_transaksi': 0,
                            'terakhir_transaksi': t.get('tanggal_transaksi', '-')
                        }
                        customer_purchases[customer_id] = []
                    
                    customer_stats[customer_id]['total_belanja'] += safe_float(t.get('total_bayar', 0))
                    customer_stats[customer_id]['total_transaksi'] += 1
                    
                    # Update terakhir transaksi jika lebih baru
                    tgl_trans = t.get('tanggal_transaksi', '-')
                    if tgl_trans > customer_stats[customer_id]['terakhir_transaksi']:
                        customer_stats[customer_id]['terakhir_transaksi'] = tgl_trans
                    
                    # Ambil detail items dari transaksi
                    items = t.get('items', [])
                    if not items:
                        items = t.get('details', [])
                    
                    for item in items:
                        customer_purchases[customer_id].append({
                            'tanggal': t.get('tanggal_transaksi', '-')[:19],
                            'no_invoice': t.get('no_invoice', '-'),
                            'brand': item.get('brand', '-'),
                            'jumlah_karung': safe_float(item.get('jumlah_karung', 0)),
                            'kg': safe_float(item.get('jumlah_kg', 0)),
                            'subtotal': safe_float(item.get('subtotal', 0))
                        })
            
            # Gabungkan dengan data pelanggan
            customer_list = []
            for c in self.customers:
                customer_id = c.get('id_pelanggan')
                stats = customer_stats.get(customer_id, {
                    'total_belanja': 0,
                    'total_transaksi': 0,
                    'terakhir_transaksi': '-'
                })
                
                customer_list.append({
                    'id_pelanggan': customer_id,
                    'kode_pelanggan': c.get('kode_pelanggan', '-'),
                    'nama': c.get('nama', '-'),
                    'total_belanja': stats['total_belanja'],
                    'total_transaksi': stats['total_transaksi'],
                    'terakhir_transaksi': stats['terakhir_transaksi'][:19] if stats['terakhir_transaksi'] != '-' else '-',
                    'purchases': customer_purchases.get(customer_id, [])
                })
            
            # Sort by total belanja
            customer_list.sort(key=lambda x: x['total_belanja'], reverse=True)
            
            # Simpan untuk filtering
            self.all_customer_data = customer_list
            self.display_customer_report(customer_list)
            
        except Exception as e:
            print(f"Error updating customer tab: {e}")
            self.customer_table.setRowCount(0)

    def display_customer_report(self, customers):
        """Display customer report in table"""
        print(f"Displaying {len(customers)} customers")  # Debug
        
        self.customer_table.setRowCount(len(customers))
        self.customer_table.setSelectionMode(QTableWidget.SingleSelection)
        
        for i, c in enumerate(customers):
            self.customer_table.setItem(i, 0, QTableWidgetItem(c.get('kode_pelanggan', '-')))
            self.customer_table.setItem(i, 1, QTableWidgetItem(c.get('nama', '-')))
            self.customer_table.setItem(i, 2, QTableWidgetItem(format_rupiah(c.get('total_belanja', 0))))
            self.customer_table.setItem(i, 3, QTableWidgetItem(str(c.get('total_transaksi', 0))))
        
        # Alternating row colors
        for i in range(self.customer_table.rowCount()):
            for j in range(self.customer_table.columnCount()):
                item = self.customer_table.item(i, j)
                if item and i % 2 == 0:
                    item.setBackground(QColor(248, 250, 252))
        
        # Resize rows to contents
        self.customer_table.resizeRowsToContents()

    def filter_customer_report(self):
        """Filter customer report by search text"""
        if not hasattr(self, 'all_customer_data'):
            return
        
        search_text = self.customer_search_input.text().lower()
        
        if not search_text:
            self.display_customer_report(self.all_customer_data)
            # Pilih baris pertama
            if self.all_customer_data:
                self.customer_table.selectRow(0)
            return
        
        filtered = [c for c in self.all_customer_data 
                    if search_text in c.get('nama', '').lower() 
                    or search_text in c.get('kode_pelanggan', '').lower()]
        
        self.display_customer_report(filtered)
        
        # Pilih baris pertama jika ada
        if filtered:
            self.customer_table.selectRow(0)
        else:
            self.selected_customer_name.setText("Tidak ada pelanggan")
            self.customer_detail_table.setRowCount(1)
            self.customer_detail_table.setItem(0, 0, QTableWidgetItem("Tidak ada data"))

    def on_customer_table_double_click(self, item):
        """Handle double click on customer table - show full transaction history"""
        row = item.row()
        if hasattr(self, 'all_customer_data') and row < len(self.all_customer_data):
            customer_data = self.all_customer_data[row]
            
            # Ambil data lengkap customer dari database
            customer_id = customer_data.get('id_pelanggan')
            if customer_id:
                customer = DatabaseHelper.get_customer_by_id(customer_id)
                if customer:
                    from windows.customer_history_dialog import CustomerHistoryDialog
                    dialog = CustomerHistoryDialog(customer, self)
                    dialog.exec()

    def show_customer_purchase_details(self, customer):
        """Show purchase details for selected customer"""
        purchases = customer.get('purchases', [])
        
        if not purchases:
            self.customer_detail_table.setRowCount(1)
            self.customer_detail_table.setItem(0, 0, QTableWidgetItem("Belum ada transaksi"))
            # Kosongkan kolom lain
            for col in range(1, 5):
                if self.customer_detail_table.columnCount() > col:
                    self.customer_detail_table.setItem(0, col, QTableWidgetItem(""))
            return
        
        # Sort by tanggal (most recent first)
        purchases.sort(key=lambda x: x.get('tanggal', ''), reverse=True)
        
        self.customer_detail_table.setRowCount(len(purchases))
        
        for i, p in enumerate(purchases):
            self.customer_detail_table.setItem(i, 0, QTableWidgetItem(p.get('tanggal', '-')[:16]))
            self.customer_detail_table.setItem(i, 1, QTableWidgetItem(p.get('no_invoice', '-')))
            self.customer_detail_table.setItem(i, 2, QTableWidgetItem(p.get('brand', '-')))
            self.customer_detail_table.setItem(i, 3, QTableWidgetItem(f"{p.get('jumlah_karung', 0):.1f}"))
            self.customer_detail_table.setItem(i, 4, QTableWidgetItem(format_rupiah(p.get('subtotal', 0))))
        
        # Alternating row colors
        for i in range(self.customer_detail_table.rowCount()):
            for j in range(self.customer_detail_table.columnCount()):
                item = self.customer_detail_table.item(i, j)
                if item and i % 2 == 0:
                    item.setBackground(QColor(248, 250, 252))
        
        # Resize rows
        self.customer_detail_table.resizeRowsToContents()
    
    def on_period_change(self, period):
        """Handle period change"""
        if period == "Kustom":
            self.start_date.setVisible(True)
            self.end_date.setVisible(True)
            self.current_periode = 'kustom'
        else:
            self.start_date.setVisible(False)
            self.end_date.setVisible(False)
            if period == "Harian":
                self.current_periode = 'harian'
            elif period == "Bulanan":
                self.current_periode = 'bulanan'
            else:
                self.current_periode = 'tahunan'
        self.load_report()
    
    def load_report(self):
        """Load all report data"""
        try:
            # Clear summary cards
            for i in reversed(range(self.summary_layout.count())):
                w = self.summary_layout.itemAt(i).widget()
                if w:
                    w.deleteLater()
            
            # Get data from database with details
            if self.current_periode == 'kustom':
                start = self.start_date.date().toString("yyyy-MM-dd")
                end = self.end_date.date().toString("yyyy-MM-dd")
                self.transactions = DatabaseHelper.get_transactions_with_details(start_date=start, end_date=end)
            else:
                self.transactions = DatabaseHelper.get_transactions_with_details()
            
            # Ensure transactions is a list
            if self.transactions is None:
                self.transactions = []
            
            self.filtered_transactions = self.filter_transactions_by_period(self.transactions)
            
            # Calculate summary
            summary = self.calculate_summary(self.filtered_transactions)
            
            # Display summary cards
            self.display_summary_cards(summary)
            
            # Update all tabs
            self.update_transaction_tab(self.filtered_transactions)
            self.update_brand_tab(self.filtered_transactions)
            self.update_customer_tab()
            self.update_stock_tab()
            
        except Exception as e:
            print(f"Error loading report: {e}")
            traceback.print_exc()
    
    def filter_transactions_by_period(self, transactions):
        """Filter transactions based on selected period"""
        if not transactions:
            return []
            
        if self.current_periode == 'harian':
            today = datetime.now().strftime("%Y-%m-%d")
            return [t for t in transactions if t.get('tanggal_transaksi', '')[:10] == today]
        elif self.current_periode == 'bulanan':
            this_month = datetime.now().strftime("%Y-%m")
            return [t for t in transactions if t.get('tanggal_transaksi', '')[:7] == this_month]
        elif self.current_periode == 'tahunan':
            this_year = datetime.now().strftime("%Y")
            return [t for t in transactions if t.get('tanggal_transaksi', '')[:4] == this_year]
        else:
            return transactions
    
    def calculate_summary(self, transactions):
        """Calculate summary statistics"""
        if not transactions:
            return {
                'total_penjualan': 0,
                'total_transaksi': 0,
                'total_karung': 0,
                'total_kg': 0,
                'rata_rata': 0,
                'tertinggi': 0,
                'terendah': 0
            }
        
        total_penjualan = sum(safe_float(t.get('total_bayar', 0)) for t in transactions)
        total_transaksi = len(transactions)
        total_karung = sum(safe_float(t.get('total_karung', 0)) for t in transactions)
        total_kg = sum(safe_float(t.get('total_kg', 0)) for t in transactions)
        
        all_totals = [safe_float(t.get('total_bayar', 0)) for t in transactions]
        
        return {
            'total_penjualan': total_penjualan,
            'total_transaksi': total_transaksi,
            'total_karung': total_karung,
            'total_kg': total_kg,
            'rata_rata': total_penjualan / total_transaksi if total_transaksi > 0 else 0,
            'tertinggi': max(all_totals) if all_totals else 0,
            'terendah': min(all_totals) if all_totals else 0
        }
    
    def display_summary_cards(self, summary):
        """Display summary cards"""
        cards = [
            ("💰 Total Penjualan", format_rupiah(summary.get('total_penjualan', 0)), "#10b981"),
            ("🔄 Total Transaksi", str(summary.get('total_transaksi', 0)), "#3b82f6"),
            ("📦 Total Karung", f"{summary.get('total_karung', 0):.1f}", "#f59e0b"),
            ("⚖️ Total KG", f"{summary.get('total_kg', 0):.0f}", "#ef4444"),
            ("📊 Rata-rata Transaksi", format_rupiah(summary.get('rata_rata', 0)), "#8b5cf6"),
            ("📈 Tertinggi", format_rupiah(summary.get('tertinggi', 0)), "#ec489a"),
            ("📉 Terendah", format_rupiah(summary.get('terendah', 0)), "#6b7280")
        ]
        
        for title, value, color in cards:
            card = QFrame()
            card.setFixedHeight(150)
            card.setStyleSheet("""
                QFrame {
                    background-color: white;
                    border-radius: 10px;
                    padding: 10px;
                    border: 1px solid #e2e8f0;
                }
            """)
            card_layout = QVBoxLayout()
            title_label = QLabel(title)
            title_label.setStyleSheet("color: #64748b; font-size: 11px;")
            value_label = QLabel(value)
            value_label.setStyleSheet(f"font-size: 18px; font-weight: 600; color: {color};")
            card_layout.addWidget(title_label)
            card_layout.addWidget(value_label)
            card.setLayout(card_layout)
            self.summary_layout.addWidget(card)
    
    def update_transaction_tab(self, transactions):
        """Update transaction table with customer data"""
        if not transactions:
            self.trans_table.setRowCount(0)
            return
        
        # Sort by date, most recent first
        sorted_trans = sorted(transactions, key=lambda x: x.get('tanggal_transaksi', ''), reverse=True)
        
        self.trans_table.setRowCount(len(sorted_trans))
        
        for i, t in enumerate(sorted_trans):
            # Ambil nama pelanggan
            customer_name = "-"
            customer_id = t.get('id_pelanggan')
            if customer_id:
                # Coba ambil dari data transaksi
                if 'pelanggan' in t and t['pelanggan']:
                    customer_name = t['pelanggan'].get('nama', '-')
                else:
                    # Ambil dari database
                    customer = DatabaseHelper.get_customer_by_id(customer_id)
                    if customer:
                        customer_name = customer.get('nama', '-')
            
            self.trans_table.setItem(i, 0, QTableWidgetItem(str(t.get('no_invoice', '-'))))
            self.trans_table.setItem(i, 1, QTableWidgetItem(str(t.get('tanggal_transaksi', '-'))[:19]))
            self.trans_table.setItem(i, 2, QTableWidgetItem(str(t.get('kasir', '-'))))
            self.trans_table.setItem(i, 3, QTableWidgetItem(customer_name))
            self.trans_table.setItem(i, 4, QTableWidgetItem(f"{safe_float(t.get('total_karung', 0)):.1f}"))
            self.trans_table.setItem(i, 5, QTableWidgetItem(f"{safe_float(t.get('total_kg', 0)):.0f}"))
            self.trans_table.setItem(i, 6, QTableWidgetItem(format_rupiah(t.get('total_bayar', 0))))
    
    def update_brand_tab(self, transactions):
        """Update brand sales table from transaction items"""
        if not transactions:
            self.brand_table.setRowCount(0)
            self.brand_table.setHorizontalHeaderLabels(["Brand", "Transaksi", "Karung", "KG", "Total Penjualan"])
            return
        
        # Aggregate by brand from transaction items
        brand_data = {}
        
        for t in transactions:
            # Ambil items dari transaksi (sudah diambil dari API)
            items = t.get('items', [])
            
            # Debug: print items untuk memastikan ada data
            if items:
                print(f"Transaction {t.get('no_invoice')} has {len(items)} items")
            
            for item in items:
                brand = item.get('brand', 'Unknown')
                qty = safe_float(item.get('jumlah_karung', 0))
                subtotal = safe_float(item.get('subtotal', 0))
                
                if brand not in brand_data:
                    brand_data[brand] = {
                        'qty': 0,
                        'subtotal': 0,
                        'count': 0,
                        'transactions': set()
                    }
                
                brand_data[brand]['qty'] += qty
                brand_data[brand]['subtotal'] += subtotal
                brand_data[brand]['count'] += 1
                brand_data[brand]['transactions'].add(t.get('no_invoice'))
        
        # Convert to list
        brand_list = []
        for brand, data in brand_data.items():
            if brand and brand != 'Unknown':
                brand_list.append({
                    'brand': brand,
                    'count': len(data['transactions']),  # Jumlah transaksi yang mengandung brand ini
                    'items_count': data['count'],  # Jumlah item (bisa lebih dari transaksi)
                    'qty': data['qty'],
                    'kg': data['qty'] * 50,
                    'subtotal': data['subtotal']
                })
        
        # Sort by total sales
        brand_list.sort(key=lambda x: x['subtotal'], reverse=True)
        
        # Debug: print brand data
        print(f"Brand data: {len(brand_list)} brands found")
        for b in brand_list[:5]:
            print(f"  {b['brand']}: {b['qty']} karung, Rp {b['subtotal']:,.0f}")
        
        self.brand_table.setRowCount(len(brand_list))
        
        for i, b in enumerate(brand_list):
            self.brand_table.setItem(i, 0, QTableWidgetItem(b['brand']))
            self.brand_table.setItem(i, 1, QTableWidgetItem(str(b['count'])))
            self.brand_table.setItem(i, 2, QTableWidgetItem(f"{b['qty']:.1f}"))
            self.brand_table.setItem(i, 3, QTableWidgetItem(f"{b['kg']:.0f}"))
            self.brand_table.setItem(i, 4, QTableWidgetItem(format_rupiah(b['subtotal'])))
    
    def update_customer_tab(self):
        """Update customer report table with purchase details"""
        try:
            self.customers = DatabaseHelper.get_customers()
            
            if not self.customers:
                self.customer_table.setRowCount(0)
                self.customer_table.setHorizontalHeaderLabels(["Kode", "Nama", "Total Belanja", "Transaksi"])
                return
            
            # Ambil semua transaksi untuk menghitung detail per pelanggan
            all_transactions = DatabaseHelper.get_transactions_with_details()
            
            # Hitung statistik per pelanggan
            customer_stats = {}
            customer_purchases = {}  # Menyimpan detail pembelian per pelanggan
            
            for t in all_transactions:
                customer_id = t.get('id_pelanggan')
                if customer_id:
                    if customer_id not in customer_stats:
                        customer_stats[customer_id] = {
                            'total_belanja': 0,
                            'total_transaksi': 0,
                            'terakhir_transaksi': t.get('tanggal_transaksi', '-')
                        }
                        customer_purchases[customer_id] = []
                    
                    customer_stats[customer_id]['total_belanja'] += safe_float(t.get('total_bayar', 0))
                    customer_stats[customer_id]['total_transaksi'] += 1
                    
                    # Update terakhir transaksi jika lebih baru
                    tgl_trans = t.get('tanggal_transaksi', '-')
                    if tgl_trans > customer_stats[customer_id]['terakhir_transaksi']:
                        customer_stats[customer_id]['terakhir_transaksi'] = tgl_trans
                    
                    # Ambil detail items dari transaksi
                    items = t.get('items', [])
                    if not items:
                        items = t.get('details', [])
                    
                    for item in items:
                        customer_purchases[customer_id].append({
                            'tanggal': t.get('tanggal_transaksi', '-')[:19],
                            'no_invoice': t.get('no_invoice', '-'),
                            'brand': item.get('brand', '-'),
                            'jumlah_karung': safe_float(item.get('jumlah_karung', 0)),
                            'kg': safe_float(item.get('jumlah_kg', 0)),
                            'subtotal': safe_float(item.get('subtotal', 0))
                        })
            
            # Gabungkan dengan data pelanggan
            customer_list = []
            for c in self.customers:
                customer_id = c.get('id_pelanggan')
                stats = customer_stats.get(customer_id, {
                    'total_belanja': 0,
                    'total_transaksi': 0,
                    'terakhir_transaksi': '-'
                })
                
                customer_list.append({
                    'id_pelanggan': customer_id,
                    'kode_pelanggan': c.get('kode_pelanggan', '-'),
                    'nama': c.get('nama', '-'),
                    'total_belanja': stats['total_belanja'],
                    'total_transaksi': stats['total_transaksi'],
                    'terakhir_transaksi': stats['terakhir_transaksi'][:19] if stats['terakhir_transaksi'] != '-' else '-',
                    'purchases': customer_purchases.get(customer_id, [])
                })
            
            # Sort by total belanja
            customer_list.sort(key=lambda x: x['total_belanja'], reverse=True)
            
            # Simpan untuk filtering
            self.all_customer_data = customer_list
            
            # TAMPILKAN DATA KE TABEL
            self.display_customer_report(customer_list)
            
            # Jika ada data, pilih baris pertama
            if customer_list:
                self.customer_table.selectRow(0)
            else:
                self.selected_customer_name.setText("Tidak ada pelanggan")
                self.customer_detail_table.setRowCount(1)
                self.customer_detail_table.setItem(0, 0, QTableWidgetItem("Tidak ada data"))
            
        except Exception as e:
            print(f"Error updating customer tab: {e}")
            import traceback
            traceback.print_exc()
            self.customer_table.setRowCount(0)

    def update_stock_tab(self):
        """Update stock report table with statistics"""
        try:
            self.products = DatabaseHelper.get_products()
            
            if not self.products:
                self.stock_table.setRowCount(0)
                return
            
            # Get stock logs summary
            stock_logs = DatabaseHelper.get_stock_logs()
            
            # Create summary for each product
            product_summary = {}
            for log in stock_logs:
                product_id = log.get('id_barang')
                if product_id not in product_summary:
                    product_summary[product_id] = {'masuk': 0, 'keluar': 0}
                if log.get('jenis_transaksi') == 'MASUK':
                    product_summary[product_id]['masuk'] += safe_float(log.get('jumlah_karung', 0))
                else:
                    product_summary[product_id]['keluar'] += safe_float(log.get('jumlah_karung', 0))
            
            # Sort by stock
            sorted_products = sorted(self.products, key=lambda x: safe_float(x.get('stok_karung', 0)))
            
            # Add columns for movement summary
            self.stock_table.setColumnCount(8)
            self.stock_table.setHorizontalHeaderLabels([
                "Kode", "Brand", "Stok (Karung)", "Stok (KG)", 
                "Harga/KG", "Stok Masuk", "Stok Keluar", "Status"
            ])
            
            self.stock_table.setRowCount(len(sorted_products))
            
            for i, p in enumerate(sorted_products):
                product_id = p.get('id_barang')
                masuk = product_summary.get(product_id, {}).get('masuk', 0)
                keluar = product_summary.get(product_id, {}).get('keluar', 0)
                
                stok = safe_float(p.get('stok_karung', 0))
                stok_min = safe_float(p.get('stok_minimum_karung', 2))
                stok_kg = safe_float(p.get('stok_kg', 0))
                harga = safe_float(p.get('harga_jual_kg', 0))
                
                if stok <= 0:
                    status = "Habis"
                    status_color = "#ef4444"
                elif stok <= stok_min:
                    status = "Stok Menipis"
                    status_color = "#f59e0b"
                else:
                    status = "Stok Aman"
                    status_color = "#10b981"
                
                self.stock_table.setItem(i, 0, QTableWidgetItem(p.get('kode_barang', '-')))
                self.stock_table.setItem(i, 1, QTableWidgetItem(p.get('brand', '-')))
                self.stock_table.setItem(i, 2, QTableWidgetItem(f"{stok:.1f}"))
                self.stock_table.setItem(i, 3, QTableWidgetItem(f"{stok_kg:.0f}"))
                self.stock_table.setItem(i, 4, QTableWidgetItem(format_rupiah(harga)))
                self.stock_table.setItem(i, 5, QTableWidgetItem(f"{masuk:.1f}"))
                self.stock_table.setItem(i, 6, QTableWidgetItem(f"{keluar:.1f}"))
                
                status_item = QTableWidgetItem(status)
                status_item.setForeground(QColor(status_color))
                self.stock_table.setItem(i, 7, status_item)
                
        except Exception as e:
            print(f"Error updating stock tab: {e}")
            self.stock_table.setRowCount(0)
    
    # pages/report_page.py - Update method export_to_excel

    def export_to_excel(self):
        """Export report to Excel with multiple sheets - Lengkap dengan history pelanggan"""
        try:
            # Tanyakan lokasi penyimpanan
            filename, _ = QFileDialog.getSaveFileName(
                self, "Simpan Laporan Excel", 
                f"laporan_{self.current_periode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not filename:
                return
            
            # Buat workbook
            wb = Workbook()
            
            # Hapus sheet default
            wb.remove(wb.active)
            
            # Style untuk header
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            title_font = Font(bold=True, size=14)
            subtitle_font = Font(italic=True, size=11)
            
            # ==================== SHEET 1: RINGKASAN ====================
            ws_summary = wb.create_sheet("Ringkasan")
            
            # Title
            ws_summary['A1'] = f"LAPORAN {self.period_combo.currentText().upper()}"
            ws_summary['A1'].font = title_font
            ws_summary.merge_cells('A1:D1')
            
            ws_summary['A2'] = f"Periode: {self.get_period_text()}"
            ws_summary['A2'].font = subtitle_font
            ws_summary.merge_cells('A2:D2')
            
            ws_summary['A3'] = f"Tanggal Export: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            ws_summary['A3'].font = subtitle_font
            ws_summary.merge_cells('A3:D3')
            
            # Ringkasan Penjualan
            row = 5
            ws_summary[f'A{row}'] = "RINGKASAN PENJUALAN"
            ws_summary[f'A{row}'].font = Font(bold=True, size=12, color="10b981")
            ws_summary.merge_cells(f'A{row}:D{row}')
            row += 1
            
            summary = self.calculate_summary(self.filtered_transactions)
            summary_data = [
                ("Total Penjualan", format_rupiah(summary.get('total_penjualan', 0))),
                ("Total Transaksi", str(summary.get('total_transaksi', 0))),
                ("Total Karung", f"{summary.get('total_karung', 0):.1f}"),
                ("Total KG", f"{summary.get('total_kg', 0):.0f}"),
                ("Rata-rata Transaksi", format_rupiah(summary.get('rata_rata', 0))),
                ("Transaksi Tertinggi", format_rupiah(summary.get('tertinggi', 0))),
                ("Transaksi Terendah", format_rupiah(summary.get('terendah', 0))),
            ]
            
            for label, value in summary_data:
                ws_summary[f'A{row}'] = label
                ws_summary[f'A{row}'].font = Font(bold=True)
                ws_summary[f'B{row}'] = value
                row += 1
            
            # Ringkasan Stok
            row += 1
            ws_summary[f'A{row}'] = "RINGKASAN STOK"
            ws_summary[f'A{row}'].font = Font(bold=True, size=12, color="10b981")
            ws_summary.merge_cells(f'A{row}:D{row}')
            row += 1
            
            products = DatabaseHelper.get_products()
            total_stok = sum(safe_float(p.get('stok_karung', 0)) for p in products)
            total_kg = sum(safe_float(p.get('stok_kg', 0)) for p in products)
            stok_menipis = sum(1 for p in products if safe_float(p.get('stok_karung', 0)) <= safe_float(p.get('stok_minimum_karung', 2)))
            
            stock_summary = [
                ("Total Brand", str(len(products))),
                ("Total Stok", f"{total_stok:.1f} karung"),
                ("Total KG", f"{total_kg:.0f} KG"),
                ("Stok Menipis", f"{stok_menipis} brand"),
            ]
            
            for label, value in stock_summary:
                ws_summary[f'A{row}'] = label
                ws_summary[f'A{row}'].font = Font(bold=True)
                ws_summary[f'B{row}'] = value
                row += 1
            
            # ==================== SHEET 2: DETAIL TRANSAKSI ====================
            ws_trans = wb.create_sheet("Detail Transaksi")
            
            # Header
            headers = ["No Invoice", "Tanggal", "Kasir", "Pelanggan", "Karung", "KG", "Total"]
            for col, header in enumerate(headers, 1):
                cell = ws_trans.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Data
            for row_idx, t in enumerate(self.filtered_transactions, 2):
                # Ambil nama pelanggan
                customer_name = "-"
                customer_id = t.get('id_pelanggan')
                if customer_id:
                    customer = DatabaseHelper.get_customer_by_id(customer_id)
                    if customer:
                        customer_name = customer.get('nama', '-')
                
                ws_trans.cell(row=row_idx, column=1, value=t.get('no_invoice', '-'))
                ws_trans.cell(row=row_idx, column=2, value=t.get('tanggal_transaksi', '-')[:19])
                ws_trans.cell(row=row_idx, column=3, value=t.get('kasir', '-'))
                ws_trans.cell(row=row_idx, column=4, value=customer_name)
                ws_trans.cell(row=row_idx, column=5, value=float(t.get('total_karung', 0)))
                ws_trans.cell(row=row_idx, column=6, value=float(t.get('total_kg', 0)))
                ws_trans.cell(row=row_idx, column=7, value=float(t.get('total_bayar', 0)))
            
            # Auto fit columns
            for col in range(1, 8):
                ws_trans.column_dimensions[get_column_letter(col)].width = 15
            
            # ==================== SHEET 3: PENJUALAN PER BRAND ====================
            ws_brand = wb.create_sheet("Penjualan per Brand")
            
            # Header
            headers = ["Brand", "No Invoice", "Tanggal", "Pelanggan", "Jumlah (Karung)", "KG", "Subtotal"]
            for col, header in enumerate(headers, 1):
                cell = ws_brand.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Data
            row_idx = 2
            for t in self.filtered_transactions:
                items = t.get('items', [])
                if not items:
                    items = t.get('details', [])
                
                for item in items:
                    brand = item.get('brand', 'Unknown')
                    qty = safe_float(item.get('jumlah_karung', 0))
                    subtotal = safe_float(item.get('subtotal', 0))
                    
                    # Ambil nama pelanggan
                    customer_name = "-"
                    customer_id = t.get('id_pelanggan')
                    if customer_id:
                        customer = DatabaseHelper.get_customer_by_id(customer_id)
                        if customer:
                            customer_name = customer.get('nama', '-')
                    
                    ws_brand.cell(row=row_idx, column=1, value=brand)
                    ws_brand.cell(row=row_idx, column=2, value=t.get('no_invoice', '-'))
                    ws_brand.cell(row=row_idx, column=3, value=t.get('tanggal_transaksi', '-')[:19])
                    ws_brand.cell(row=row_idx, column=4, value=customer_name)
                    ws_brand.cell(row=row_idx, column=5, value=float(qty))
                    ws_brand.cell(row=row_idx, column=6, value=float(qty * 50))
                    ws_brand.cell(row=row_idx, column=7, value=float(subtotal))
                    row_idx += 1
            
            # Auto fit columns
            for col in range(1, 8):
                ws_brand.column_dimensions[get_column_letter(col)].width = 18
            
            # ==================== SHEET 4: LOG STOK ====================
            ws_log = wb.create_sheet("Log Stok")
            
            headers = ["Tanggal", "Jenis", "Kode Barang", "Brand", "Jumlah (Karung)", "Jumlah (KG)", 
                    "Stok Sebelum (Karung)", "Stok Sesudah (Karung)", "Keterangan", "User"]
            for col, header in enumerate(headers, 1):
                cell = ws_log.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Get stock logs sesuai periode
            if self.current_periode == 'kustom':
                start = self.start_date.date().toString("yyyy-MM-dd")
                end = self.end_date.date().toString("yyyy-MM-dd")
                logs = DatabaseHelper.get_stock_logs(start_date=start, end_date=end)
            elif self.current_periode == 'harian':
                date = self.start_date.date().toString("yyyy-MM-dd") if self.start_date.isVisible() else datetime.now().strftime("%Y-%m-%d")
                logs = DatabaseHelper.get_stock_logs(start_date=date, end_date=date)
            elif self.current_periode == 'bulanan':
                year = self.start_date.date().year() if self.start_date.isVisible() else datetime.now().year
                month = self.start_date.date().month() if self.start_date.isVisible() else datetime.now().month
                start = f"{year}-{month:02d}-01"
                last_day = calendar.monthrange(year, month)[1]
                end = f"{year}-{month:02d}-{last_day}"
                logs = DatabaseHelper.get_stock_logs(start_date=start, end_date=end)
            elif self.current_periode == 'tahunan':
                year = self.start_date.date().year() if self.start_date.isVisible() else datetime.now().year
                start = f"{year}-01-01"
                end = f"{year}-12-31"
                logs = DatabaseHelper.get_stock_logs(start_date=start, end_date=end)
            else:
                logs = DatabaseHelper.get_stock_logs()
            
            for row_idx, log in enumerate(logs, 2):
                ws_log.cell(row=row_idx, column=1, value=log.get('created_at', '-')[:19])
                ws_log.cell(row=row_idx, column=2, value=log.get('jenis_transaksi', '-'))
                ws_log.cell(row=row_idx, column=3, value=log.get('kode_barang', '-'))
                ws_log.cell(row=row_idx, column=4, value=log.get('brand', '-'))
                ws_log.cell(row=row_idx, column=5, value=float(log.get('jumlah_karung', 0)))
                ws_log.cell(row=row_idx, column=6, value=float(log.get('jumlah_kg', 0)))
                ws_log.cell(row=row_idx, column=7, value=float(log.get('stok_sebelum_karung', 0)))
                ws_log.cell(row=row_idx, column=8, value=float(log.get('stok_sesudah_karung', 0)))
                ws_log.cell(row=row_idx, column=9, value=log.get('keterangan', '-'))
                ws_log.cell(row=row_idx, column=10, value=log.get('user', '-'))
                
                # Warna untuk jenis transaksi
                jenis = log.get('jenis_transaksi', '-')
                if jenis == 'MASUK':
                    for col in range(1, 11):
                        ws_log.cell(row=row_idx, column=col).fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                elif jenis == 'KELUAR':
                    for col in range(1, 11):
                        ws_log.cell(row=row_idx, column=col).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            
            for col in range(1, 11):
                ws_log.column_dimensions[get_column_letter(col)].width = 18
            
            # ==================== SHEET 5: REKAP BRAND ====================
            ws_brand_summary = wb.create_sheet("Rekap Brand")
            
            headers = ["Brand", "Jumlah Transaksi", "Total Karung", "Total KG", "Total Penjualan"]
            for col, header in enumerate(headers, 1):
                cell = ws_brand_summary.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            brand_summary = {}
            for t in self.filtered_transactions:
                items = t.get('items', [])
                if not items:
                    items = t.get('details', [])
                
                for item in items:
                    brand = item.get('brand', 'Unknown')
                    qty = safe_float(item.get('jumlah_karung', 0))
                    subtotal = safe_float(item.get('subtotal', 0))
                    
                    if brand not in brand_summary:
                        brand_summary[brand] = {'qty': 0, 'subtotal': 0, 'transactions': set()}
                    
                    brand_summary[brand]['qty'] += qty
                    brand_summary[brand]['subtotal'] += subtotal
                    brand_summary[brand]['transactions'].add(t.get('no_invoice'))
            
            sorted_brands = sorted(brand_summary.items(), key=lambda x: x[1]['subtotal'], reverse=True)
            
            for row_idx, (brand, data) in enumerate(sorted_brands, 2):
                ws_brand_summary.cell(row=row_idx, column=1, value=brand)
                ws_brand_summary.cell(row=row_idx, column=2, value=len(data['transactions']))
                ws_brand_summary.cell(row=row_idx, column=3, value=float(data['qty']))
                ws_brand_summary.cell(row=row_idx, column=4, value=float(data['qty'] * 50))
                ws_brand_summary.cell(row=row_idx, column=5, value=float(data['subtotal']))
            
            for col in range(1, 6):
                ws_brand_summary.column_dimensions[get_column_letter(col)].width = 18
            
            # ==================== SHEET 6: LAPORAN PELANGGAN (DETAIL) ====================
            ws_customer = wb.create_sheet("Laporan Pelanggan")
            
            # Header
            headers = ["Kode Pelanggan", "Nama", "No Telepon", "Email", "Total Belanja", "Total Transaksi", "Poin"]
            for col, header in enumerate(headers, 1):
                cell = ws_customer.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Data pelanggan
            customers = DatabaseHelper.get_customers()
            customer_data = []
            for c in customers:
                customer_id = c.get('id_pelanggan')
                # Hitung total transaksi dari database
                trans_count = 0
                for t in self.filtered_transactions:
                    if t.get('id_pelanggan') == customer_id:
                        trans_count += 1
                
                customer_data.append({
                    'kode': c.get('kode_pelanggan', '-'),
                    'nama': c.get('nama', '-'),
                    'telp': c.get('no_telp', '-'),
                    'email': c.get('email', '-'),
                    'total_belanja': safe_float(c.get('total_belanja', 0)),
                    'transaksi': trans_count,
                    'poin': safe_float(c.get('poin', 0))
                })
            
            # Sort by total belanja
            customer_data.sort(key=lambda x: x['total_belanja'], reverse=True)
            
            for row_idx, c in enumerate(customer_data, 2):
                ws_customer.cell(row=row_idx, column=1, value=c['kode'])
                ws_customer.cell(row=row_idx, column=2, value=c['nama'])
                ws_customer.cell(row=row_idx, column=3, value=c['telp'])
                ws_customer.cell(row=row_idx, column=4, value=c['email'])
                ws_customer.cell(row=row_idx, column=5, value=float(c['total_belanja']))
                ws_customer.cell(row=row_idx, column=6, value=c['transaksi'])
                ws_customer.cell(row=row_idx, column=7, value=float(c['poin']))
            
            for col in range(1, 8):
                ws_customer.column_dimensions[get_column_letter(col)].width = 18
            
            # ==================== SHEET 7: HISTORY PELANGGAN ====================
            ws_customer_history = wb.create_sheet("History Pelanggan")
            
            # Header
            headers = ["Pelanggan", "No Invoice", "Tanggal", "Brand", "Jumlah (Karung)", "KG", "Subtotal"]
            for col, header in enumerate(headers, 1):
                cell = ws_customer_history.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Data history per pelanggan
            row_idx = 2
            for c in customers:
                customer_id = c.get('id_pelanggan')
                customer_name = c.get('nama', '-')
                
                for t in self.filtered_transactions:
                    if t.get('id_pelanggan') == customer_id:
                        items = t.get('items', [])
                        if not items:
                            items = t.get('details', [])
                        
                        for item in items:
                            brand = item.get('brand', 'Unknown')
                            qty = safe_float(item.get('jumlah_karung', 0))
                            subtotal = safe_float(item.get('subtotal', 0))
                            
                            ws_customer_history.cell(row=row_idx, column=1, value=customer_name)
                            ws_customer_history.cell(row=row_idx, column=2, value=t.get('no_invoice', '-'))
                            ws_customer_history.cell(row=row_idx, column=3, value=t.get('tanggal_transaksi', '-')[:19])
                            ws_customer_history.cell(row=row_idx, column=4, value=brand)
                            ws_customer_history.cell(row=row_idx, column=5, value=float(qty))
                            ws_customer_history.cell(row=row_idx, column=6, value=float(qty * 50))
                            ws_customer_history.cell(row=row_idx, column=7, value=float(subtotal))
                            row_idx += 1
            
            for col in range(1, 8):
                ws_customer_history.column_dimensions[get_column_letter(col)].width = 18
            
            # Simpan file
            wb.save(filename)
            
            msg_box = QMessageBox()
            msg_box.setWindowTitle("Sukses!")
            msg_box.setText(f"✅ Laporan berhasil diexport ke:\n{filename}\n\nFile berisi 7 sheet:\n"
                        "1. Ringkasan\n2. Detail Transaksi\n3. Penjualan per Brand\n4. Log Stok\n"
                        "5. Rekap Brand\n6. Laporan Pelanggan\n7. History Pelanggan")
            msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg_box.setDefaultButton(QMessageBox.StandardButton.Ok)
            msg_box.setIcon(QMessageBox.Icon.Information)
            msg_box.exec()
            
        except ImportError:
            QMessageBox.warning(self, "Error", "❌ Modul openpyxl tidak terinstall!\n\nJalankan: pip install openpyxl")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"❌ Gagal export: {e}")
            import traceback
            traceback.print_exc()

    def get_period_text(self):
        """Get period text for report"""
        if self.current_periode == 'harian':
            date = self.start_date.date().toString("dd/MM/yyyy") if self.start_date.isVisible() else datetime.now().strftime("%d/%m/%Y")
            return f"{date}"
        elif self.current_periode == 'bulanan':
            month = self.start_date.date().month() if self.start_date.isVisible() else datetime.now().month
            year = self.start_date.date().year() if self.start_date.isVisible() else datetime.now().year
            return f"{month:02d}/{year}"
        elif self.current_periode == 'tahunan':
            year = self.start_date.date().year() if self.start_date.isVisible() else datetime.now().year
            return f"{year}"
        else:
            start = self.start_date.date().toString("dd/MM/yyyy")
            end = self.end_date.date().toString("dd/MM/yyyy")
            return f"{start} - {end}"

    def create_stock_log_tab(self):
        """Tab untuk log pemasukan dan pengeluaran stok"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Filter bar untuk log stok
        filter_layout = QHBoxLayout()
        
        # Filter jenis
        filter_layout.addWidget(QLabel("Jenis:"))
        self.log_type_combo = QComboBox()
        self.log_type_combo.addItems(["Semua", "MASUK", "KELUAR", "TRANSAKSI"])
        self.log_type_combo.setMinimumHeight(32)
        self.log_type_combo.setFixedWidth(100)
        self.log_type_combo.currentTextChanged.connect(self.filter_stock_logs)
        filter_layout.addWidget(self.log_type_combo)
        
        filter_layout.addSpacing(15)
        
        # Cari produk
        filter_layout.addWidget(QLabel("Cari:"))
        self.log_search_input = QLineEdit()
        self.log_search_input.setPlaceholderText("Cari brand atau kode...")
        self.log_search_input.setMinimumHeight(32)
        self.log_search_input.setFixedWidth(200)
        self.log_search_input.textChanged.connect(self.filter_stock_logs)
        filter_layout.addWidget(self.log_search_input)
        
        filter_layout.addStretch()
        
        # Tombol refresh log
        refresh_log_btn = QPushButton("🔄 Refresh")
        refresh_log_btn.setMinimumHeight(32)
        refresh_log_btn.setFixedWidth(100)
        refresh_log_btn.setStyleSheet(STYLE_REFRESH_BUTTON)
        refresh_log_btn.clicked.connect(self.load_stock_logs)
        filter_layout.addWidget(refresh_log_btn)
        
        layout.addLayout(filter_layout)
        
        # Tabel Log Stok
        self.log_table = QTableWidget()
        self.log_table.setColumnCount(8)
        self.log_table.setHorizontalHeaderLabels([
            "Tanggal", "Jenis", "Kode", "Brand", "Jumlah (KG)",
            "Stok Sebelum (Karung)", "Jumlah (Karung)", "Stok Sesudah (Karung)"
        ])
        self.log_table.horizontalHeader().setStretchLastSection(True)
        self.log_table.setAlternatingRowColors(True)
        
        # Set column widths
        self.log_table.setColumnWidth(0, 180)  # Tanggal
        self.log_table.setColumnWidth(1, 100)   # Jenis
        self.log_table.setColumnWidth(2, 100)   # Kode
        self.log_table.setColumnWidth(3, 150)  # Brand
        self.log_table.setColumnWidth(4, 120)   # Jumlah KG
        self.log_table.setColumnWidth(5, 150)  # Stok Sebelum
        self.log_table.setColumnWidth(6, 150)  # Jumlah Karung
        self.log_table.setColumnWidth(7, 150)  # Stok Sesudah
        
        self.log_table.verticalHeader().setDefaultSectionSize(40)
        self.log_table.verticalHeader().setVisible(False)
        
        layout.addWidget(self.log_table)
        
        # Info
        info_label = QLabel("💡 Menampilkan semua perubahan stok (pemasukan dan pengeluaran)")
        info_label.setStyleSheet("color: #64748b; font-size: 11px; margin-top: 8px;")
        layout.addWidget(info_label)
        
        widget.setLayout(layout)
        return widget
    
    def load_stock_logs(self):
        """Load stock logs from database"""
        try:
            # Get date range from filter
            if self.current_periode == 'kustom':
                start = self.start_date.date().toString("yyyy-MM-dd")
                end = self.end_date.date().toString("yyyy-MM-dd")
                self.stock_logs = DatabaseHelper.get_stock_logs(start_date=start, end_date=end)
            elif self.current_periode == 'harian':
                date = self.start_date.date().toString("yyyy-MM-dd") if self.start_date.isVisible() else datetime.now().strftime("%Y-%m-%d")
                self.stock_logs = DatabaseHelper.get_stock_logs(start_date=date, end_date=date)
            elif self.current_periode == 'bulanan':
                year = self.start_date.date().year() if self.start_date.isVisible() else datetime.now().year
                month = self.start_date.date().month() if self.start_date.isVisible() else datetime.now().month
                start = f"{year}-{month:02d}-01"
                last_day = calendar.monthrange(year, month)[1]
                end = f"{year}-{month:02d}-{last_day}"
                self.stock_logs = DatabaseHelper.get_stock_logs(start_date=start, end_date=end)
            elif self.current_periode == 'tahunan':
                year = self.start_date.date().year() if self.start_date.isVisible() else datetime.now().year
                start = f"{year}-01-01"
                end = f"{year}-12-31"
                self.stock_logs = DatabaseHelper.get_stock_logs(start_date=start, end_date=end)
            else:
                self.stock_logs = DatabaseHelper.get_stock_logs()
            
            self.display_stock_logs(self.stock_logs)
        except Exception as e:
            print(f"Error loading stock logs: {e}")
    
    def display_stock_logs(self, logs):
        """Display stock logs in table"""
        if not logs:
            self.log_table.setRowCount(0)
            return
        
        self.log_table.setRowCount(len(logs))
        
        for i, log in enumerate(logs):
            # Tanggal
            tanggal = log.get('created_at', '-')[:19]
            self.log_table.setItem(i, 0, QTableWidgetItem(tanggal))
            
            # Jenis dengan warna
            jenis = log.get('jenis_transaksi', '-')
            jenis_item = QTableWidgetItem(jenis)
            if jenis == "MASUK":
                jenis_item.setForeground(QColor(16, 185, 129))  # Hijau
                jenis_item.setBackground(QColor(220, 252, 231))
            else:
                jenis_item.setForeground(QColor(239, 68, 68))   # Merah
                jenis_item.setBackground(QColor(254, 226, 226))
            self.log_table.setItem(i, 1, jenis_item)
            
            # Kode
            self.log_table.setItem(i, 2, QTableWidgetItem(log.get('kode_barang', '-')))
            
            # Brand
            self.log_table.setItem(i, 3, QTableWidgetItem(log.get('brand', '-')))
            
            # Jumlah Karung
            jumlah_karung = safe_float(log.get('jumlah_karung', 0))
            self.log_table.setItem(i, 6, QTableWidgetItem(f"{jumlah_karung:.1f}"))
            
            # Jumlah KG
            jumlah_kg = safe_float(log.get('jumlah_kg', 0))
            self.log_table.setItem(i, 4, QTableWidgetItem(f"{jumlah_kg:.0f}"))
            
            # Stok Sebelum
            stok_sebelum = safe_float(log.get('stok_sebelum_karung', 0))
            self.log_table.setItem(i, 5, QTableWidgetItem(f"{stok_sebelum:.1f}"))
            
            # Stok Sesudah
            stok_sesudah = safe_float(log.get('stok_sesudah_karung', 0))
            self.log_table.setItem(i, 7, QTableWidgetItem(f"{stok_sesudah:.1f}"))
    
    def filter_stock_logs(self):
        """Filter stock logs by type and search"""
        if not hasattr(self, 'stock_logs'):
            return
        
        filtered = self.stock_logs
        
        # Filter berdasarkan jenis
        jenis = self.log_type_combo.currentText()
        if jenis != "Semua":
            filtered = [l for l in filtered if l.get('jenis_transaksi') == jenis]
        
        # Filter berdasarkan pencarian
        search = self.log_search_input.text().lower()
        if search:
            filtered = [l for l in filtered if 
                       search in l.get('brand', '').lower() or 
                       search in l.get('kode_barang', '').lower()]
        
        self.display_stock_logs(filtered)
    
    def load_report(self):
        """Load all report data"""
        try:
            # Clear summary cards
            for i in reversed(range(self.summary_layout.count())):
                w = self.summary_layout.itemAt(i).widget()
                if w:
                    w.deleteLater()
            
            # Get data from database
            if self.current_periode == 'kustom':
                start = self.start_date.date().toString("yyyy-MM-dd")
                end = self.end_date.date().toString("yyyy-MM-dd")
                self.transactions = DatabaseHelper.get_transactions_with_details(start_date=start, end_date=end)
            else:
                self.transactions = DatabaseHelper.get_transactions_with_details()
            
            # Ensure transactions is a list
            if self.transactions is None:
                self.transactions = []
            
            # Filter transactions by period
            self.filtered_transactions = self.filter_transactions_by_period(self.transactions)
            
            # Calculate summary
            summary = self.calculate_summary(self.filtered_transactions)
            
            # Display summary cards
            self.display_summary_cards(summary)
            
            # Update all tabs
            self.update_transaction_tab(self.filtered_transactions)
            self.update_brand_tab(self.filtered_transactions)
            self.update_customer_tab()
            self.update_stock_tab()
            
            # Load stock logs
            self.load_stock_logs()
            
        except Exception as e:
            print(f"Error loading report: {e}")
            traceback.print_exc()